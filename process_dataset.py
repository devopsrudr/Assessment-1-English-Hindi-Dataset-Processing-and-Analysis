"""
process_dataset.py
-------------------
Assignment 1: English-Hindi Dataset Processing and Analysis

Reads the parallel corpus (data/eng.txt, data/hin.txt) from the
`ainlpml/english-hindi` Hugging Face dataset, computes word counts,
applies the required filters, and writes a cleaned Excel workbook.

Filters applied (as per assignment spec):
  1. Keep only sentence pairs where BOTH the English and Hindi word count
     fall within [5, 50].
  2. Of those, keep only pairs where the word-count difference
     (English count - Hindi count) falls within [-10, +10].

Usage:
    python process_dataset.py
    (reads from ./data/eng.txt and ./data/hin.txt by default)

Output:
    output/cleaned_english_hindi_dataset.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---- Configuration ----
# All paths are resolved relative to this script's own location, so the
# script works no matter what directory it's run from.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)

ENG_FILE = os.path.join(PROJECT_ROOT, "data", "eng.txt")
HIN_FILE = os.path.join(PROJECT_ROOT, "data", "hin.txt")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "cleaned_english_hindi_dataset.xlsx")

MIN_WORDS = 5
MAX_WORDS = 50
MIN_DIFF = -10
MAX_DIFF = 10


def word_count(sentence: str) -> int:
    """Word count = number of whitespace-separated tokens.
    Works for both English and Hindi (Devanagari script also uses
    spaces to separate words)."""
    return len(sentence.split())


def load_parallel_sentences(eng_path: str, hin_path: str):
    with open(eng_path, encoding="utf-8") as f:
        eng_lines = [line.strip() for line in f.read().splitlines()]
    with open(hin_path, encoding="utf-8") as f:
        hin_lines = [line.strip() for line in f.read().splitlines()]

    if len(eng_lines) != len(hin_lines):
        raise ValueError(
            f"Line count mismatch: {len(eng_lines)} English lines vs "
            f"{len(hin_lines)} Hindi lines. Files must be aligned 1:1."
        )
    return eng_lines, hin_lines


def filter_pairs(eng_lines, hin_lines):
    """Apply word-count-range and word-count-difference filters."""
    kept = []
    for eng, hin in zip(eng_lines, hin_lines):
        if not eng or not hin:
            continue
        we = word_count(eng)
        wh = word_count(hin)

        # Filter 1: both word counts must fall within [MIN_WORDS, MAX_WORDS]
        if not (MIN_WORDS <= we <= MAX_WORDS):
            continue
        if not (MIN_WORDS <= wh <= MAX_WORDS):
            continue

        # Filter 2: word count difference must fall within [MIN_DIFF, MAX_DIFF]
        diff = we - wh
        if not (MIN_DIFF <= diff <= MAX_DIFF):
            continue

        kept.append((eng, hin, we, wh, diff))

    return kept


def write_excel(rows, output_path: str):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Dataset"

    headers = [
        "English Sentence",
        "Hindi Sentence",
        "Word Count (English)",
        "Word Count (Hindi)",
        "Difference (Eng - Hindi)",
    ]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="Arial", size=11)

    for r, (eng, hin, _we, _wh, _diff) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=eng).font = body_font
        ws.cell(row=r, column=2, value=hin).font = body_font
        # Word counts and difference as LIVE FORMULAS (whitespace-token count),
        # so the sheet recalculates automatically if a sentence is edited.
        ws.cell(
            row=r, column=3,
            value=f'=IF(TRIM(A{r})="",0,LEN(TRIM(A{r}))-LEN(SUBSTITUTE(TRIM(A{r})," ",""))+1)'
        ).font = body_font
        ws.cell(
            row=r, column=4,
            value=f'=IF(TRIM(B{r})="",0,LEN(TRIM(B{r}))-LEN(SUBSTITUTE(TRIM(B{r})," ",""))+1)'
        ).font = body_font
        ws.cell(row=r, column=5, value=f"=C{r}-D{r}").font = body_font

    # Column widths
    widths = [70, 70, 20, 20, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(output_path)


def main():
    print("Loading parallel sentences...")
    eng_lines, hin_lines = load_parallel_sentences(ENG_FILE, HIN_FILE)
    print(f"  Loaded {len(eng_lines)} sentence pairs.")

    print("Applying word-count range and difference filters...")
    rows = filter_pairs(eng_lines, hin_lines)
    print(f"  {len(rows)} sentence pairs remain after filtering "
          f"(word count {MIN_WORDS}-{MAX_WORDS} in both languages, "
          f"difference {MIN_DIFF} to {MAX_DIFF}).")

    print(f"Writing Excel output to {OUTPUT_FILE} ...")
    write_excel(rows, OUTPUT_FILE)
    print("Done.")
    print(
        "\nNOTE: Word count / difference columns are Excel formulas "
        "(auto-recalculate). Run scripts/recalc equivalent or open in "
        "Excel/LibreOffice once to populate cached values."
    )


if __name__ == "__main__":
    main()
