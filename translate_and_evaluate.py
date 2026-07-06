"""
translate_and_evaluate.py
--------------------------
Assignment 2: Translation with an LLM (local, open-source Hugging Face model)

Steps performed:
  1. Load the cleaned dataset produced by Assignment 1
     (../assignment1_dataset_processing/output/cleaned_english_hindi_dataset.xlsx)
  2. Take the first N (default 100) English sentences from it.
  3. Translate each one into Hindi using a local Hugging Face MT model
     (default: Helsinki-NLP/opus-mt-en-hi).
  4. Compare model translations against the dataset's own Hindi sentences
     (used as reference translations) using BLEU, CHRF, and TER
     (via the `sacrebleu` library).
  5. Save:
       - output/translations.xlsx   (English sentence, Model-generated Hindi)
       - output/scores.txt          (BLEU / CHRF / TER results)

Usage:
    python translate_and_evaluate.py
    python translate_and_evaluate.py --num-sentences 100 --model Helsinki-NLP/opus-mt-en-hi

Notes:
  - First run will download the model from Hugging Face (a few hundred MB).
    Make sure you have internet access and, if the model is gated, a valid
    HF_TOKEN (see the .env file in this folder, same pattern as the
    Assignment 1 dataset downloader).
  - Runs fine on CPU for 100 sentences with opus-mt-en-hi (small model).
    If you have a GPU, it will be used automatically.
"""

import argparse
import os
import time

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT = os.path.join(
    SCRIPT_DIR, "..", "assignment1_dataset_processing", "output",
    "cleaned_english_hindi_dataset.xlsx",
)
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
OUTPUT_XLSX = os.path.join(OUTPUT_DIR, "translations.xlsx")
OUTPUT_SCORES = os.path.join(OUTPUT_DIR, "scores.txt")


def load_sentences(input_path: str, num_sentences: int):
    """Load the cleaned Assignment-1 dataset and take the first N rows."""
    df = pd.read_excel(input_path)
    df = df.head(num_sentences)
    english_sentences = df["English Sentence"].tolist()
    reference_hindi = df["Hindi Sentence"].tolist()
    return english_sentences, reference_hindi


def translate_sentences(english_sentences, model_name: str, batch_size: int = 8):
    """Translate a list of English sentences into Hindi using a local
    Hugging Face translation model."""
    from transformers import pipeline
    import torch

    device = 0 if torch.cuda.is_available() else -1
    print(f"Loading model '{model_name}' (device={'GPU' if device == 0 else 'CPU'})...")
    translator = pipeline("translation", model=model_name, device=device)

    translations = []
    start = time.time()
    for i in range(0, len(english_sentences), batch_size):
        batch = english_sentences[i:i + batch_size]
        results = translator(batch, max_length=256)
        translations.extend([r["translation_text"] for r in results])
        print(f"  Translated {min(i + batch_size, len(english_sentences))}"
              f"/{len(english_sentences)} sentences...")
    elapsed = time.time() - start
    print(f"Translation finished in {elapsed:.1f}s.")
    return translations


def compute_scores(hypotheses, references):
    """Compute corpus-level BLEU, CHRF, and TER using sacrebleu."""
    import sacrebleu

    bleu = sacrebleu.corpus_bleu(hypotheses, [references])
    chrf = sacrebleu.corpus_chrf(hypotheses, [references])
    ter = sacrebleu.corpus_ter(hypotheses, [references])

    return {
        "BLEU": bleu.score,
        "CHRF": chrf.score,
        "TER": ter.score,
    }


def write_excel(english_sentences, model_translations, output_path: str):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Translations"

    headers = ["Original English Sentence", "Model-generated Hindi Translation"]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="Arial", size=11)
    for r, (eng, hin) in enumerate(zip(english_sentences, model_translations), start=2):
        ws.cell(row=r, column=1, value=eng).font = body_font
        ws.cell(row=r, column=2, value=hin).font = body_font

    ws.column_dimensions[get_column_letter(1)].width = 70
    ws.column_dimensions[get_column_letter(2)].width = 70
    ws.freeze_panes = "A2"
    wb.save(output_path)


def write_scores(scores: dict, model_name: str, num_sentences: int, output_path: str):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("Assignment 2 - Translation Evaluation Scores\n")
        f.write("=" * 50 + "\n")
        f.write(f"Model used: {model_name}\n")
        f.write(f"Number of sentences evaluated: {num_sentences}\n")
        f.write(
            "Reference translations: Hindi sentences from the Assignment 1 "
            "cleaned dataset (ainlpml/english-hindi)\n"
        )
        f.write("-" * 50 + "\n")
        for metric, value in scores.items():
            f.write(f"{metric}: {value:.4f}\n")
        f.write("-" * 50 + "\n")
        f.write(
            "Notes:\n"
            "  - BLEU and CHRF: higher is better (0-100 scale).\n"
            "  - TER (Translation Edit Rate): lower is better "
            "(0 = perfect match, can exceed 100 for very poor output).\n"
        )


def main():
    parser = argparse.ArgumentParser(description="Assignment 2: LLM Translation + Evaluation")
    parser.add_argument("--input", default=DEFAULT_INPUT,
                         help="Path to Assignment 1 cleaned Excel file")
    parser.add_argument("--num-sentences", type=int, default=100,
                         help="Number of English sentences to translate")
    parser.add_argument("--model", default="Helsinki-NLP/opus-mt-en-hi",
                         help="Hugging Face model name/id for translation")
    parser.add_argument("--batch-size", type=int, default=8)
    args = parser.parse_args()

    print(f"Loading {args.num_sentences} sentences from {args.input} ...")
    english_sentences, reference_hindi = load_sentences(args.input, args.num_sentences)
    print(f"  Loaded {len(english_sentences)} sentences.")

    model_translations = translate_sentences(
        english_sentences, args.model, args.batch_size
    )

    print("Computing BLEU / CHRF / TER scores...")
    scores = compute_scores(model_translations, reference_hindi)
    for metric, value in scores.items():
        print(f"  {metric}: {value:.4f}")

    print(f"Writing Excel output to {OUTPUT_XLSX} ...")
    write_excel(english_sentences, model_translations, OUTPUT_XLSX)

    print(f"Writing scores to {OUTPUT_SCORES} ...")
    write_scores(scores, args.model, len(english_sentences), OUTPUT_SCORES)

    print("\nDone. Files created:")
    print(f"  - {OUTPUT_XLSX}")
    print(f"  - {OUTPUT_SCORES}")


if __name__ == "__main__":
    main()
